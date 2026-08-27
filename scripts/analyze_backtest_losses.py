from collections import defaultdict
from datetime import timedelta

from openpyxl import load_workbook


path = "reports/PHLC_XAUUSD_M15_90_Days_Backtest.xlsx"
sheet = load_workbook(path, data_only=True, read_only=True)["Trades"]
headers = [cell.value for cell in next(sheet.iter_rows())]
trades = [dict(zip(headers, (cell.value for cell in row))) for row in sheet.iter_rows()]
closed = [trade for trade in trades if trade["Result"] in {"TP", "SL"}]


def summarize(key):
    groups = defaultdict(list)
    for trade in closed:
        groups[key(trade)].append(trade)
    rows = []
    for name, items in groups.items():
        losses = sum(item["Result"] == "SL" for item in items)
        rows.append((name, len(items), losses, losses / len(items), sum(item["P&L"] for item in items)))
    return sorted(rows, key=lambda row: (-row[3], -row[2]))


print("IST_HOUR")
for row in summarize(lambda t: (t["Entry Time UTC"] + timedelta(hours=5, minutes=30)).hour):
    if row[1] >= 10:
        print(row)

print("WEEKDAY")
for row in summarize(lambda t: t["Entry Time UTC"].strftime("%A")):
    print(row)

print("IST_BLOCK")
for row in summarize(lambda t: f"{((t['Entry Time UTC']+timedelta(hours=5,minutes=30)).hour//3)*3:02d}:00-{(((t['Entry Time UTC']+timedelta(hours=5,minutes=30)).hour//3)*3)+3:02d}:00"):
    print(row)

def sl_bin(trade):
    distance = trade["Entry"] - trade["Stop Loss"]
    if distance < 2: return "<2"
    if distance < 4: return "2-4"
    if distance < 6: return "4-6"
    if distance < 10: return "6-10"
    return "10+"

print("SL_DISTANCE")
for row in summarize(sl_bin):
    print(row)

print("WORST_DATES")
for row in sorted(summarize(lambda t: t["Entry Time UTC"].date()), key=lambda value: (value[4], value[2]))[:10]:
    print(row)
