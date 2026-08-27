from pathlib import Path
from datetime import datetime, timezone, timedelta

from openpyxl import load_workbook
from openpyxl import Workbook


args = __import__("sys").argv
timeframe = "M1" if "--m1" in args else "H4" if "--h4" in args else "H1" if "--h1" in args else "M15"
target_suffix = "_TP30" if "--target30" in args else "_TP5" if "--target5" in args else ""
side_suffix = "_SELL" if "--sell" in args else ""
filename = f"PHLC_XAUUSD_{timeframe}_90_Days_Backtest{target_suffix}{side_suffix}.xlsx"
report = Path(__file__).resolve().parents[1] / "reports" / filename
ws = load_workbook(report, read_only=True, data_only=True)["Trades"]
rows = list(ws.iter_rows(values_only=True))
headers = {name: i for i, name in enumerate(rows[0])}
ist = timezone(timedelta(hours=5, minutes=30))

trades = []
for row in rows[1:]:
    outcome = row[headers["Result"]]
    if outcome not in {"TP", "SL"}:
        continue
    entry = row[headers["Entry Time UTC"]]
    if isinstance(entry, str):
        entry = datetime.fromisoformat(entry.replace("Z", "+00:00"))
    if entry.tzinfo is None:
        entry = entry.replace(tzinfo=timezone.utc)
    if "--june" in args and entry.month != 6:
        continue
    if "--june" not in args and entry.month == 7:
        continue
    trades.append((entry.astimezone(ist), outcome, float(row[headers["P&L"]] or 0)))


def stats(label, predicate):
    selected = [t for t in trades if predicate(t[0])]
    wins = sum(t[1] == "TP" for t in selected)
    losses = sum(t[1] == "SL" for t in selected)
    pnl = sum(t[2] for t in selected)
    rate = wins / len(selected) * 100 if selected else 0
    print(f"{label}|{len(selected)}|{wins}|{losses}|{rate:.2f}|{pnl:.2f}")


stats("ALL", lambda dt: True)
stats("09-24 IST", lambda dt: 9 <= dt.hour < 24)
stats("21-24 IST", lambda dt: 21 <= dt.hour < 24)
stats("00-09 IST", lambda dt: 0 <= dt.hour < 9)
stats("06:30-22:30 IST", lambda dt: 390 <= dt.hour * 60 + dt.minute <= 1350)

if "--june" in args:
    output = report.with_name(report.stem + "_JUNE_ONLY.xlsx")
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["SELL M15 - JUNE ONLY", "10-point target / previous candle high SL"])
    summary.append(["All June trades", len(trades)])
    summary.append(["Wins", sum(t[1] == "TP" for t in trades)])
    summary.append(["Losses", sum(t[1] == "SL" for t in trades)])
    summary.append(["Net P&L", sum(t[2] for t in trades)])
    details = wb.create_sheet("Trades")
    details.append(rows[0])
    for row in rows[1:]:
        entry = row[headers["Entry Time UTC"]]
        if isinstance(entry, str):
            entry = datetime.fromisoformat(entry.replace("Z", "+00:00"))
        if entry and entry.month == 6:
            details.append(row)
    details.freeze_panes = "A2"
    details.auto_filter.ref = details.dimensions
    wb.save(output)
    print(f"REPORT|{output}")
