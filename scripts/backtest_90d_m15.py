from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path
import sys

import MetaTrader5 as mt5
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


SYMBOL = "XAUUSD"
TARGET_DISTANCE = 30.0 if "--target30" in sys.argv else 5.0 if "--target5" in sys.argv else 10.0
LOT = 0.01
MAGIC = 26082026
SIDE = "SELL" if "--sell" in sys.argv else "BUY"
if "--m1" in sys.argv:
    TIMEFRAME, TIMEFRAME_NAME = mt5.TIMEFRAME_M1, "M1"
elif "--h4" in sys.argv:
    TIMEFRAME, TIMEFRAME_NAME = mt5.TIMEFRAME_H4, "H4"
elif "--h1" in sys.argv:
    TIMEFRAME, TIMEFRAME_NAME = mt5.TIMEFRAME_H1, "H1"
else:
    TIMEFRAME, TIMEFRAME_NAME = mt5.TIMEFRAME_M15, "M15"
target_suffix = f"_TP{int(TARGET_DISTANCE)}" if TARGET_DISTANCE != 10.0 else ""
side_suffix = "_SELL" if SIDE == "SELL" else ""
OUTPUT = Path(f"reports/PHLC_XAUUSD_{TIMEFRAME_NAME}_90_Days_Backtest{target_suffix}{side_suffix}.xlsx")


def main() -> None:
    if not mt5.initialize():
        raise RuntimeError(f"MT5 initialize failed: {mt5.last_error()}")
    try:
        tick = mt5.symbol_info_tick(SYMBOL)
        spec = mt5.symbol_info(SYMBOL)
        if tick is None or spec is None:
            raise RuntimeError(f"Symbol unavailable: {SYMBOL}")
        end = datetime.fromtimestamp(int(tick.time), UTC)
        start = end - timedelta(days=90)
        if TIMEFRAME_NAME == "M1":
            rate_chunks = []
            chunk_start = start
            while chunk_start < end:
                chunk_end = min(chunk_start + timedelta(days=10), end)
                chunk = mt5.copy_rates_range(SYMBOL, TIMEFRAME, chunk_start, chunk_end)
                if chunk is not None and len(chunk):
                    rate_chunks.extend(chunk)
                chunk_start = chunk_end
            rates = rate_chunks
        else:
            rates = mt5.copy_rates_range(SYMBOL, TIMEFRAME, start, end)
        if rates is None or len(rates) < 3:
            raise RuntimeError(f"Insufficient {TIMEFRAME_NAME} history: {mt5.last_error()}")

        bars = [
            {
                "time": datetime.fromtimestamp(int(rate["time"]), UTC),
                "open": float(rate["open"]), "high": float(rate["high"]),
                "low": float(rate["low"]), "close": float(rate["close"]),
                "spread": float(rate["spread"]) * float(spec.point),
            }
            for rate in rates
        ]
        trades, signals = run_backtest(bars)
        write_excel(trades, signals, bars, start, end)
        print(f"bars={len(bars)} trades={len(trades)} output={OUTPUT.resolve()}")
    finally:
        mt5.shutdown()


def run_backtest(bars: list[dict]) -> tuple[list[dict], list[dict]]:
    reference = None
    armed = True
    pending = None
    position = None
    trades: list[dict] = []
    signals: list[dict] = []

    for bar in bars:
        if pending and position is None:
            entry = bar["open"] + bar["spread"] if SIDE == "BUY" else bar["open"]
            position = {
                **pending,
                "entry_time": bar["time"], "entry": entry,
                "sl": pending["reference_low"] if SIDE == "BUY" else pending["reference_high"],
                "tp": entry + TARGET_DISTANCE if SIDE == "BUY" else entry - TARGET_DISTANCE,
            }
            pending = None

        if position:
            sl_hit = bar["low"] <= position["sl"] if SIDE == "BUY" else bar["high"] >= position["sl"]
            tp_hit = bar["high"] >= position["tp"] if SIDE == "BUY" else bar["low"] <= position["tp"]
            if sl_hit or tp_hit:
                # OHLC cannot reveal which was first; conservative SL-first policy.
                result = "SL" if sl_hit else "TP"
                exit_price = position["sl"] if sl_hit else position["tp"]
                points = (exit_price - position["entry"]) if SIDE == "BUY" else (position["entry"] - exit_price)
                profit = mt5.order_calc_profit(
                    mt5.ORDER_TYPE_BUY if SIDE == "BUY" else mt5.ORDER_TYPE_SELL,
                    SYMBOL, LOT, position["entry"], exit_price
                )
                trades.append({
                    **position, "exit_time": bar["time"], "exit": exit_price,
                    "result": result, "points": points,
                    "profit": float(profit or 0.0),
                })
                position = None

        bullish = bar["close"] > bar["open"]
        bearish = bar["close"] < bar["open"]
        if (SIDE == "BUY" and bearish) or (SIDE == "SELL" and bullish):
            armed = True

        valid = bool(armed and reference and (
            (SIDE == "BUY" and bullish and bar["close"] > reference["high"] and bar["low"] >= reference["low"])
            or (SIDE == "SELL" and bearish and bar["close"] < reference["low"] and bar["high"] <= reference["high"])
        ))
        if valid:
            armed = False
            accepted = position is None and pending is None
            signals.append({
                "time": bar["time"], "reference_time": reference["time"],
                "reference_high": reference["high"], "reference_low": reference["low"],
                "signal_open": bar["open"], "signal_high": bar["high"],
                "signal_low": bar["low"], "signal_close": bar["close"],
                "status": "Accepted" if accepted else "Skipped - active trade",
            })
            if accepted:
                pending = {
                    "signal_time": bar["time"], "reference_time": reference["time"],
                    "reference_high": reference["high"], "reference_low": reference["low"],
                    "signal_open": bar["open"], "signal_high": bar["high"],
                    "signal_low": bar["low"], "signal_close": bar["close"],
                }
        if ((SIDE == "BUY" and bullish) or (SIDE == "SELL" and bearish)) and armed:
            reference = bar
        elif valid:
            reference = bar

    if position:
        last = bars[-1]
        exit_price = last["close"]
        profit = mt5.order_calc_profit(mt5.ORDER_TYPE_BUY if SIDE == "BUY" else mt5.ORDER_TYPE_SELL, SYMBOL, LOT, position["entry"], exit_price)
        trades.append({**position, "exit_time": last["time"], "exit": exit_price,
                       "result": "Open at end", "points": (exit_price-position["entry"]) if SIDE == "BUY" else (position["entry"]-exit_price),
                       "profit": float(profit or 0.0)})
    return trades, signals


def write_excel(trades: list[dict], signals: list[dict], bars: list[dict], start: datetime, end: datetime) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    wins = sum(t["result"] == "TP" for t in trades)
    losses = sum(t["result"] == "SL" for t in trades)
    net_points = sum(t["points"] for t in trades)
    net_profit = sum(t["profit"] for t in trades)
    gross_win = sum(max(t["profit"], 0) for t in trades)
    gross_loss = abs(sum(min(t["profit"], 0) for t in trades))
    rows = [
        ("PHLC BUY BACKTEST", "XAUUSD · M15 · 90 Days"),
        ("Period", f"{start:%Y-%m-%d %H:%M UTC} to {end:%Y-%m-%d %H:%M UTC}"),
        ("Bars", len(bars)), ("Trades", len(trades)), ("Wins", wins), ("Losses", losses),
        ("Win rate", wins / max(wins + losses, 1)),
        ("Net price points", net_points), ("Net P&L", net_profit),
        ("Profit factor", gross_win / gross_loss if gross_loss else None),
        ("Target", f"{TARGET_DISTANCE:.2f} direct price points"), ("Stop Loss", "Reference candle Low"),
        ("Lot", LOT), ("Reset", "Red candle required after every Buy signal"),
    ]
    for row in rows:
        summary.append(row)
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["B1"].font = Font(size=16, bold=True, color="FFFFFF")
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="17683F")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 48
    summary["B7"].number_format = "0.00%"

    trade_sheet = wb.create_sheet("Trades")
    headers = ["#", "Signal Time UTC", "Reference Time UTC", "Entry Time UTC", "Exit Time UTC",
               "Reference High", "Reference Low", "Signal O", "Signal H", "Signal L", "Signal C",
               "Entry", "Stop Loss", "Target", "Exit", "Result", "Price Points", "P&L", "Equity Curve"]
    trade_sheet.append(headers)
    equity = 0.0
    for index, trade in enumerate(trades, 1):
        equity += trade["profit"]
        trade_sheet.append([index, excel_time(trade["signal_time"]), excel_time(trade["reference_time"]), excel_time(trade["entry_time"]),
                            excel_time(trade["exit_time"]), trade["reference_high"], trade["reference_low"],
                            trade["signal_open"], trade["signal_high"], trade["signal_low"],
                            trade["signal_close"], trade["entry"], trade["sl"], trade["tp"],
                            trade["exit"], trade["result"], trade["points"], trade["profit"], equity])
    style_table(trade_sheet, len(headers))
    if trades:
        chart = LineChart()
        chart.title = "Cumulative P&L"
        chart.y_axis.title = "Account Currency"
        chart.add_data(Reference(trade_sheet, min_col=19, min_row=1, max_row=len(trades)+1), titles_from_data=True)
        chart.set_categories(Reference(trade_sheet, min_col=1, min_row=2, max_row=len(trades)+1))
        summary.add_chart(chart, "D3")

    signal_sheet = wb.create_sheet("Signals")
    signal_headers = ["#", "Signal Time UTC", "Reference Time UTC", "Reference High", "Reference Low",
                      "Signal O", "Signal H", "Signal L", "Signal C", "Status"]
    signal_sheet.append(signal_headers)
    for index, signal in enumerate(signals, 1):
        signal_sheet.append([index, excel_time(signal["time"]), excel_time(signal["reference_time"]), signal["reference_high"],
                             signal["reference_low"], signal["signal_open"], signal["signal_high"],
                             signal["signal_low"], signal["signal_close"], signal["status"]])
    style_table(signal_sheet, len(signal_headers))

    assumptions = wb.create_sheet("Methodology")
    notes = [
        "Buy-only PHLC strategy.",
        "Reference and confirmation candles must be bullish.",
        "Confirmation closes above reference High and never breaks reference Low.",
        "After a signal, another signal requires at least one bearish closed candle.",
        "Entry is next M15 candle Open plus recorded historical spread.",
        f"Target is Entry + {TARGET_DISTANCE:.2f} direct price units.",
        "Stop is reference candle Low.",
        "Only one position is allowed because the current configuration has one enabled leg.",
        "If SL and TP occur inside the same M15 candle, SL is counted first (conservative).",
        "This is an OHLC backtest, not tick-perfect execution; slippage and commissions are excluded.",
    ]
    assumptions.append(["BACKTEST METHODOLOGY"])
    for note in notes:
        assumptions.append([note])
    assumptions.column_dimensions["A"].width = 115
    assumptions["A1"].font = Font(bold=True, color="FFFFFF")
    assumptions["A1"].fill = PatternFill("solid", fgColor="17683F")
    wb.save(OUTPUT)


def style_table(sheet, columns: int) -> None:
    fill = PatternFill("solid", fgColor="207A4B")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(24, max(11, max(len(str(c.value or "")) for c in column) + 2))
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=min(5, columns)):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"


def excel_time(value: datetime) -> datetime:
    return value.astimezone(UTC).replace(tzinfo=None)


if __name__ == "__main__":
    main()
